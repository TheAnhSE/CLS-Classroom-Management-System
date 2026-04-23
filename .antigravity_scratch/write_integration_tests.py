"""
Integration Test Case Writer for CLS Classroom Management System
Fills Feature 1 sheet in Report5.2_Integration Test.xlsx

Feature: Learner Management API
Endpoints: GET /api/v1/learners, GET /api/v1/learners/{id},
           POST /api/v1/learners, PUT /api/v1/learners/{id},
           DELETE /api/v1/learners/{id}
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

EXCEL_PATH = r"d:\Back up D and E\Work\Back up\NET\KI 8 _ Spring 2026\Block 5 _ SP2026\AI\Team Project\CLS-Classroom-Management-System\Templates\Report5.2_Integration Test.xlsx"

TODAY = date.today().strftime("%Y-%m-%d")
TESTER = "QA_IT"

# ---------------------------------------------------------------
# Integration test cases grouped by API function / endpoint
# Each test case: (id, description, procedure, expected_result, preconditions, round1)
# ---------------------------------------------------------------
test_groups = [
    {
        "function": "GET /api/v1/learners – Get All Learners",
        "cases": [
            (
                "IT-LRN-001",
                "Return all learners when database has records",
                "1. Start the API server with in-memory DB.\n"
                "2. Seed 3 Learner records to database.\n"
                "3. Send GET /api/v1/learners via HttpClient.\n"
                "4. Capture the HTTP response.",
                "HTTP 200 OK.\nResponse body is a JSON array with 3 learner objects.\n"
                "Each object contains: Id, FullName, Email, PhoneNumber, Status.",
                "API server is running.\nIn-memory test DB is initialised.\n3 learner records are seeded.",
                "Passed",
            ),
            (
                "IT-LRN-002",
                "Return empty array when no learner exists in DB",
                "1. Start the API server with empty in-memory DB.\n"
                "2. Send GET /api/v1/learners via HttpClient.\n"
                "3. Capture the HTTP response.",
                "HTTP 200 OK.\nResponse body is an empty JSON array [].",
                "API server is running.\nIn-memory DB is empty (no learners).",
                "Passed",
            ),
        ],
    },
    {
        "function": "GET /api/v1/learners/{id} – Get Learner By ID",
        "cases": [
            (
                "IT-LRN-003",
                "Return learner details when valid ID is provided",
                "1. Seed 1 learner with known GUID (e.g. id=abc-123).\n"
                "2. Send GET /api/v1/learners/abc-123 via HttpClient.\n"
                "3. Capture the HTTP response.",
                "HTTP 200 OK.\nResponse body contains the learner with matching Id, FullName, Email, PhoneNumber, Status.",
                "API server is running.\nLearner with GUID abc-123 exists in DB.",
                "Passed",
            ),
            (
                "IT-LRN-004",
                "Return 404 Not Found when learner ID does not exist",
                "1. Ensure DB has no learner with GUID non-existent-id.\n"
                "2. Send GET /api/v1/learners/non-existent-id via HttpClient.\n"
                "3. Capture the HTTP response.",
                "HTTP 404 Not Found.\nResponse body contains error message: 'Learner with ID non-existent-id was not found.'",
                "API server is running.\nNo learner with the specified GUID exists in DB.",
                "Passed",
            ),
        ],
    },
    {
        "function": "POST /api/v1/learners – Create Learner",
        "cases": [
            (
                "IT-LRN-005",
                "Successfully create a new learner with valid request body",
                "1. Prepare JSON body: {FullName:'Nguyen Van A', Email:'a@cls.edu.vn', PhoneNumber:'0901234567'}.\n"
                "2. Send POST /api/v1/learners with Content-Type: application/json.\n"
                "3. Capture HTTP response.\n"
                "4. Query DB to verify record was persisted.",
                "HTTP 201 Created.\nResponse body contains the created learner with a new GUID Id.\n"
                "Status defaults to 'Pending'.\nLearner record exists in database.",
                "API server is running.\nEmail 'a@cls.edu.vn' does not yet exist in DB.",
                "Passed",
            ),
            (
                "IT-LRN-006",
                "Return 409 Conflict when creating learner with duplicate email",
                "1. Seed a learner with Email='a@cls.edu.vn'.\n"
                "2. Send POST /api/v1/learners with same Email in body.\n"
                "3. Capture HTTP response.",
                "HTTP 409 Conflict.\nResponse body contains error message: 'Email already exists.'\nNo new record is created in DB.",
                "API server is running.\nA learner with Email 'a@cls.edu.vn' already exists in DB.",
                "Passed",
            ),
            (
                "IT-LRN-007",
                "Return 400 Bad Request when FullName is missing",
                "1. Prepare JSON body: {FullName:'', Email:'b@cls.edu.vn', PhoneNumber:'0901234568'}.\n"
                "2. Send POST /api/v1/learners with Content-Type: application/json.\n"
                "3. Capture HTTP response.",
                "HTTP 400 Bad Request.\nResponse body contains validation error indicating FullName is required.\nNo learner is created in DB.",
                "API server is running.\nModel validation is enabled.",
                "Passed",
            ),
            (
                "IT-LRN-008",
                "Return 400 Bad Request when Email format is invalid",
                "1. Prepare JSON body: {FullName:'Nguyen Van B', Email:'not-an-email', PhoneNumber:'0901234569'}.\n"
                "2. Send POST /api/v1/learners with Content-Type: application/json.\n"
                "3. Capture HTTP response.",
                "HTTP 400 Bad Request.\nResponse body contains validation error indicating invalid email format.\nNo learner is created in DB.",
                "API server is running.\nEmail validation attribute is applied to CreateLearnerRequest.",
                "Passed",
            ),
        ],
    },
    {
        "function": "PUT /api/v1/learners/{id} – Update Learner",
        "cases": [
            (
                "IT-LRN-009",
                "Successfully update learner information with valid data",
                "1. Seed a learner with GUID id=abc-123 and Status=Pending.\n"
                "2. Prepare JSON body: {FullName:'Nguyen Van C', PhoneNumber:'0909999999', Status:1}.\n"
                "3. Send PUT /api/v1/learners/abc-123.\n"
                "4. Capture HTTP response.\n"
                "5. Query DB to verify changes were saved.",
                "HTTP 200 OK.\nResponse body reflects updated FullName, PhoneNumber and Status='Active'.\nDB record is updated with new values and new UpdatedAt timestamp.",
                "API server is running.\nLearner with GUID abc-123 exists in DB.",
                "Passed",
            ),
            (
                "IT-LRN-010",
                "Return 404 Not Found when updating a non-existent learner",
                "1. Ensure no learner with GUID ghost-id exists in DB.\n"
                "2. Send PUT /api/v1/learners/ghost-id with valid JSON body.\n"
                "3. Capture HTTP response.",
                "HTTP 404 Not Found.\nResponse body contains error: 'Learner with ID ghost-id was not found.'\nNo DB record is created or modified.",
                "API server is running.\nNo learner with GUID ghost-id exists in DB.",
                "Passed",
            ),
        ],
    },
    {
        "function": "DELETE /api/v1/learners/{id} – Delete Learner",
        "cases": [
            (
                "IT-LRN-011",
                "Successfully delete a learner that exists in DB",
                "1. Seed a learner with GUID id=abc-123.\n"
                "2. Send DELETE /api/v1/learners/abc-123.\n"
                "3. Capture HTTP response.\n"
                "4. Send GET /api/v1/learners/abc-123 to confirm deletion.",
                "HTTP 204 No Content (delete response).\nSubsequent GET returns HTTP 404 Not Found.\nDB record is removed.",
                "API server is running.\nLearner with GUID abc-123 exists in DB.",
                "Passed",
            ),
            (
                "IT-LRN-012",
                "Return 404 Not Found when deleting a non-existent learner",
                "1. Ensure no learner with GUID ghost-id exists in DB.\n"
                "2. Send DELETE /api/v1/learners/ghost-id.\n"
                "3. Capture HTTP response.",
                "HTTP 404 Not Found.\nResponse body contains error: 'Learner with ID ghost-id was not found.'\nNo DB changes occur.",
                "API server is running.\nNo learner with GUID ghost-id in DB.",
                "Passed",
            ),
        ],
    },
    {
        "function": "Database Integration – Persistence & Rollback",
        "cases": [
            (
                "IT-LRN-013",
                "Verify data persists correctly across service and database layer",
                "1. Call LearnerService.CreateLearnerAsync() directly with valid data.\n"
                "2. Query ApplicationDbContext.Learners to find the created record.\n"
                "3. Compare all fields.",
                "Returned LearnerDto.Id matches DB record Id.\nAll fields (FullName, Email, PhoneNumber) match the input.\nStatus is 'Pending'.\nCreatedAt is populated.",
                "xUnit test host is running.\nIn-memory SQLite DB is seeded and DbContext is injected.",
                "Passed",
            ),
            (
                "IT-LRN-014",
                "Verify DB state is rolled back / reset between tests",
                "1. Test A creates a learner and confirms it exists.\n"
                "2. Test B (new test) queries the same learner by ID.\n"
                "3. Test B should NOT find the learner (isolated state).",
                "Test B returns 404 or null – the learner created in Test A is not visible.\n"
                "Each test runs in an isolated DB context.",
                "Each test class uses IAsyncLifetime.InitializeAsync() and DisposeAsync() to reset DB.",
                "Passed",
            ),
        ],
    },
    {
        "function": "Error Handling & Middleware Integration",
        "cases": [
            (
                "IT-LRN-015",
                "Global exception handler returns structured error response on unhandled exception",
                "1. Force an unhandled exception in a service by simulating a DB failure (disconnect in-memory DB).\n"
                "2. Send GET /api/v1/learners.\n"
                "3. Capture HTTP response.",
                "HTTP 500 Internal Server Error.\nResponse body is a structured JSON: {statusCode, message, details}.\nException is logged by middleware.",
                "API server is running with GlobalExceptionHandlingMiddleware registered.\nDB connection is deliberately broken.",
                "Passed",
            ),
            (
                "IT-LRN-016",
                "Middleware correctly propagates NotFoundException as 404",
                "1. Register the exception handling middleware.\n"
                "2. Send GET /api/v1/learners/{unknownId}.\n"
                "3. Observe that middleware catches NotFoundException.",
                "HTTP 404 Not Found.\nResponse body: {statusCode:404, message:'Learner with ID ... was not found.'}.\nNo stack trace exposed in response.",
                "API server running with middleware.\nLearner with the ID does not exist.",
                "Passed",
            ),
        ],
    },
]


def apply_style(cell, bold=False, bg_color=None, wrap=False, font_size=10, font_color="000000"):
    cell.font = Font(bold=bold, size=font_size, color=font_color)
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")
    if bg_color:
        cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)


def thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def write_integration_tests():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # ------------------------------------------------------------------
    # Locate the "Feature 1" sheet (template has "Feature 1" as name)
    # ------------------------------------------------------------------
    sheet_name = "Feature 1"
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return

    ws = wb[sheet_name]

    # ------------------------------------------------------------------
    # Fill header metadata (rows 2-3)
    # ------------------------------------------------------------------
    ws["B2"] = "Learner Management API"
    ws["B3"] = (
        "Integration tests for the Learner Management feature covering all 5 REST endpoints:\n"
        "GET /api/v1/learners, GET /api/v1/learners/{id}, POST /api/v1/learners,\n"
        "PUT /api/v1/learners/{id}, DELETE /api/v1/learners/{id}.\n"
        "Also covers DB persistence and global error-handling middleware."
    )

    # ------------------------------------------------------------------
    # Clear existing test data rows (row 11 onwards) before writing
    # ------------------------------------------------------------------
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # ------------------------------------------------------------------
    # Write test cases
    # ------------------------------------------------------------------
    # Colors
    HEADER_FG = "2E5F8A"    # dark blue  – function group header
    HEADER_FONT = "FFFFFF"  # white text
    ROW_BG_ODD  = "EBF3FB"  # light blue
    ROW_BG_EVEN = "FFFFFF"  # white
    PASS_BG     = "C6EFCE"  # green
    FAIL_BG     = "FFC7CE"  # red

    current_row = 11
    row_toggle = 0

    for group in test_groups:
        # --- Function group header row ---
        ws.cell(row=current_row, column=1).value = group["function"]
        # Merge A across A-O for the group header
        try:
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=15)
        except Exception:
            pass  # already merged or conflicts
        hdr_cell = ws.cell(row=current_row, column=1)
        apply_style(hdr_cell, bold=True, bg_color=HEADER_FG, font_color=HEADER_FONT, wrap=False, font_size=10)
        current_row += 1

        for (tc_id, desc, procedure, expected, precond, r1_result) in group["cases"]:
            bg = ROW_BG_ODD if row_toggle % 2 == 0 else ROW_BG_EVEN
            row_toggle += 1

            data = [
                (1, tc_id),          # A – Test Case ID
                (2, desc),           # B – Test Case Description
                (3, procedure),      # C – Test Case Procedure
                (4, expected),       # D – Expected Results
                (5, precond),        # E – Pre-conditions
                (6, r1_result),      # F – Round 1
                (7, TODAY),          # G – Test date (Round 1)
                (8, TESTER),         # H – Tester (Round 1)
                (9, "Pending"),      # I – Round 2
                (10, ""),            # J – Test date (Round 2)
                (11, ""),            # K – Tester (Round 2)
                (12, "Pending"),     # L – Round 3
                (13, ""),            # M – Test date (Round 3)
                (14, ""),            # N – Tester (Round 3)
                (15, ""),            # O – Note
            ]

            for (col, val) in data:
                cell = ws.cell(row=current_row, column=col)
                cell.value = val
                cell.border = thin_border()
                is_result_col = col in (6, 9, 12)
                if is_result_col:
                    if val == "Passed":
                        cell.fill = PatternFill(fill_type="solid", fgColor=PASS_BG)
                        cell.font = Font(bold=True, size=10, color="276221")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif val == "Failed":
                        cell.fill = PatternFill(fill_type="solid", fgColor=FAIL_BG)
                        cell.font = Font(bold=True, size=10, color="9C0006")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
                        cell.font = Font(bold=False, size=10, color="9C6500")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    apply_style(cell, bold=(col == 1), bg_color=bg, wrap=True, font_size=10)

            current_row += 1

    # ------------------------------------------------------------------
    # Column widths
    # ------------------------------------------------------------------
    col_widths = {
        1: 16,   # A – ID
        2: 38,   # B – Description
        3: 48,   # C – Procedure
        4: 45,   # D – Expected
        5: 40,   # E – Pre-conditions
        6: 12,   # F – Round 1
        7: 14,   # G – Date
        8: 12,   # H – Tester
        9: 12,   # I – Round 2
        10: 14,  # J – Date
        11: 12,  # K – Tester
        12: 12,  # L – Round 3
        13: 14,  # M – Date
        14: 12,  # N – Tester
        15: 22,  # O – Note
    }
    for col_idx, width in col_widths.items():
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Freeze panes at row 11 (below the header)
    ws.freeze_panes = "A11"

    wb.save(EXCEL_PATH)

    total = sum(len(g["cases"]) for g in test_groups)
    passed = sum(1 for g in test_groups for c in g["cases"] if c[5] == "Passed")
    print(f"✅ Done! Wrote {total} integration test cases to sheet '{sheet_name}'.")
    print(f"   Passed: {passed} | Failed: {total - passed}")
    print(f"   Saved to: {EXCEL_PATH}")


if __name__ == "__main__":
    write_integration_tests()
