#!/usr/bin/env python
import sys
import os

# Add path for openpyxl
try:
    from openpyxl import load_workbook
except ImportError:
    os.system('pip install openpyxl xlrd -q')
    from openpyxl import load_workbook

file_path = r'd:\Back up D and E\Work\Back up\NET\KI 8 _ Spring 2026\Block 5 _ SP2026\AI\Team Project\CLS-Classroom-Management-System\Report5.1_Unit Test.xls'

try:
    # Try loading with openpyxl
    wb = load_workbook(file_path, data_only=False)
    print('Workbook loaded successfully with openpyxl')
    print(f'Sheet names: {wb.sheetnames}')
    
    # Check if methodName1 exists
    if 'methodName1' in wb.sheetnames:
        ws = wb['methodName1']
        print(f'\nSheet "methodName1" found!')
        print(f'Max row: {ws.max_row}, Max column: {ws.max_column}')
        print('\nFirst 15 rows of data:')
        for row_idx, row in enumerate(ws.iter_rows(max_row=15, values_only=False), 1):
            print(f'Row {row_idx}:')
            for cell in row:
                if cell.value is not None or cell.coordinate in ['A1', 'B1', 'C1', 'D1', 'E1']:
                    print(f'  {cell.coordinate}: {repr(cell.value)}')
    else:
        print('\nAvailable sheets:')
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f'  - {sheet_name} (rows: {ws.max_row}, cols: {ws.max_column})')
            
except Exception as e:
    print(f'Error: {e}')
    import traceback
    traceback.print_exc()
